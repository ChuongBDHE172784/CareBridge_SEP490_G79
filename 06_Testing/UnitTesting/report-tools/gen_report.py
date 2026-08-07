# -*- coding: utf-8 -*-
"""Generate the CareBridge Unit Test report workbook from real test results.

Layout follows 06_Testing/UnitTesting/Report5.1_UnitTest_reference2.xlsx:
Guideline / Cover / MethodList / Statistics / one sheet per use case.
"""
import json, os, re, sys
from collections import OrderedDict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))   # 06_Testing/UnitTesting/report-tools
sys.path.insert(0, HERE)
from usecases import EXCLUDED_USE_CASES, REQUIREMENTS, USE_CASES  # noqa: E402

OUT = os.path.abspath(os.path.join(HERE, "..", "Report5_Unit Test_CareBridge.xlsx"))
PROJECT_NAME = "CareBridge — Nền tảng chăm sóc mẹ và bé"
PROJECT_CODE = "SEP490_G79"
DOC_CODE = "SEP490_G79_Report5_UnitTest_v1.1"
ISSUE_DATE = "2026-08-07"     # ngày chạy test — không đổi, các ô Executed Date đều dùng mốc này
REVISION_DATE = "2026-08-08"  # ngày sửa cấu trúc tài liệu
CREATOR = "Nhóm SEP490_G79"

# ---------------------------------------------------------------- styling
THIN = Side(style="thin", color="B0B0B0")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="D9E1F2")
SEC_FILL = PatternFill("solid", fgColor="F2F2F2")
TITLE_FILL = PatternFill("solid", fgColor="305496")
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
UNT_FILL = PatternFill("solid", fgColor="FFEB9C")
B_BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(vertical="top", wrap_text=True)

# ---------------------------------------------------------------- classify
BOUNDARY = re.compile(
    r"boundar|\bmax\b|maximum|\bmin\b|minimum|limit|overflow|over\d|under\d|"
    r"threshold|truncat|largest|smallest|firstpage|lastpage|\bzero\b|"
    # No leading \b on the char count: in `qExactly100Chars` the `y1` seam is not a word boundary.
    r"empty[_ ]?list|(?<!\d)\d+[ _]?chars?\b|\d+ ?dp\b|size(over|under|exactly)|"
    # `exactly` is kept only where it sits on a numeric or temporal edge (`exactly100Chars`,
    # `dueAtExactlyNow`, `exactlyTwentiethPending`). Bare `exactly` is dropped because the codebase
    # uses it overwhelmingly for count invariants — `auditLoggedExactlyOnce`, `flushes exactly once`,
    # `hasExactlyFiveValues` — which assert cardinality, not a value boundary.
    r"exactly[ _]?(?:\d|now\b|twentieth|thirtieth|tenth)|"
    # Length only counts when it is being measured against a limit, never as prose
    # ("WHO recumbent length") — hence the required companion word.
    r"(?:max|min|token|field|input|name|title|content|password)[ _]?length|"
    r"length[ _]?(?:within|over|under|exceed|bound|limit|valid)|within[ _]?bound|out[ _]?of[ _]?bound",
    re.I)
# "boundary" also names *architectural* seams (service boundary, request boundary, JSON boundary).
# Those are ordinary tests of a layer contract, not value-boundary tests, so they must not be typed
# B. Verified against a sample: without this, `serviceBoundaryRequiresGroupJourney…` and
# `…isInertAtTheRequestBoundary` were both mistyped. `exactly`, `edge`, `length` and `cap` were
# dropped from the pattern above for the same reason — they matched prose like "WHO recumbent
# length" and "posts exactly the approved four-field body".
LAYER_BOUNDARY = re.compile(
    r"(service|request|response|module|trust|layer|architectur|control|api|json|process|package|"
    r"security|domain|transport|persistence)[_ ]?boundar", re.I)
ABNORMAL = re.compile(
    r"invalid|not[_ ]?found|notfound|missing|\bnull\b|reject|denie|deny|forbidden|unauthori|"
    r"throw|error|fail|conflict|duplicate|expire|revoke|blocked|locked|never|cannot|can't|"
    r"without|unsupported|mismatch|violat|refus|abort|corrupt|malformed|stale|"
    r"(?<!\d)(?:40[0-9]|41[0-9]|422|50[0-9])(?!\d)|no[_ ]?permission|noaccess", re.I)

# camelCase boundary, plus a break before a multi-digit run so "shouldReturn403" reads
# as "Should Return 403" instead of "Should Return403".
# A trailing segment that opens with one of these words states the *precondition*, not the outcome:
# the codebase mixes `method_condition_expectation` with `method_expectation_condition`.
# The lookahead is deliberately case-sensitive — `re.I` would also let `[A-Z]` match a lowercase
# letter, which turns "onlyTitleChanged" into a false "on…" marker.
TRAILING_CONDITION = re.compile(
    r"^(?:[Ww]hen|[Ww]ith|[Ww]ithout|[Ii]f|[Gg]iven|[Uu]nless|[Ww]hile|[Dd]uring|[Aa]fter|[Bb]efore)"
    r"(?=[A-Z])")

CAMEL = re.compile(r"(?<=[a-z0-9])(?=[A-Z])|(?<=[a-zA-Z])(?=\d{2,})")


def humanise(token: str) -> str:
    token = token.replace("(", " (").strip()
    token = CAMEL.sub(" ", token)
    token = token.replace("_", " ").strip()
    return token[:1].upper() + token[1:] if token else token


def split_case(row):
    """-> (target, condition, expectation)."""
    name = row["name"]
    target = row["short"].replace("Test", "").replace(".dart", "").replace(".test", "")
    if row["layer"] == "backend":
        # Surefire names look like `method_condition_expectation`, with an optional argument list
        # and, for @ParameterizedTest, a trailing `[n]` data-set index:
        #   toProfileResponse_shouldMapCorrectRole(Role)[1]
        # Strip both decorations before splitting, then re-attach the index to the *condition* —
        # the data set is the input, not part of the method name.
        data_set = re.search(r"\[(\d+)\]\s*$", name)
        base = re.sub(r"\(.*?\)", "", name)
        base = re.sub(r"\[\d+\]\s*$", "", base).strip()
        parts = [p for p in base.split("_") if p]
        if len(parts) >= 3 and TRAILING_CONDITION.match(parts[-1]):
            # `method_expectation_whenCondition` — the tail is the precondition.
            exp = parts[-2]
            cond = "_".join(parts[1:-2] + [parts[-1]])
        elif len(parts) >= 3:
            cond, exp = "_".join(parts[1:-1]), parts[-1]
        elif len(parts) == 2:
            # `method_expectation`: the method call is the condition, the tail is the assertion.
            cond, exp = parts[0], parts[1]
        else:
            # A single camelCase sentence states only the expectation; the condition is whatever
            # the class fixture sets up, so say that rather than repeat the sentence twice.
            cond, exp = "", base
        cond = humanise(cond) if cond else "Tiền đề mặc định của test fixture"
        if data_set:
            cond = f"{cond} — bộ dữ liệu {data_set.group(1)}"
        return target, cond, humanise(exp)
    # mobile / web: the enclosing group is the scenario, the leaf is the expectation
    group = re.sub(r"\s+", " ", row.get("group") or "").strip()
    leaf = re.sub(r"\s+", " ", row.get("leaf") or name).strip()
    condition = (group or leaf)[:180] or "(unnamed)"
    expectation = leaf[:180] or "Kết quả đúng đặc tả"
    return target, condition, expectation


def case_type(row, cond, exp):
    blob = f"{row['name']} {cond} {exp}"
    if BOUNDARY.search(LAYER_BOUNDARY.sub(" ", blob)):
        return "B"
    if ABNORMAL.search(blob):
        return "A"
    return "N"


# Defects these unit tests actually caught, and that were then fixed in application code. They must
# appear in the report even though every test is green now: the test run is what surfaced them, and
# a 4.5k-case report with an empty defect list reads as if nothing was ever exercised.
# `match` is (test class / file, test-name prefix) — the case that fails if the fix is reverted.
FIXED_DEFECTS = [
    {
        "id": "DEF-RBAC-001",
        "match": ("RedFlagRuleControllerSecurityTest", "allEndpoints_asNonSystemAdmin_shouldReturn403"),
        "severity": "Cao — RBAC",
        "component": "05_Development/CareBridgeAPI/.../triage/controller/RedFlagRuleController.java",
        "symptom": "MODERATOR, CONTENT_ADMIN, EXPERT, MOTHER, FAMILY đều tạo/sửa/xoá/đọc được "
                   "red-flag rule qua /api/v1/admin/red-flag-rules (kỳ vọng 403, thực tế 200).",
        "cause": "Commit ff31c960 nới @PreAuthorize cấp class thành "
                 "hasAnyRole('SYSTEM_ADMIN','ADMIN','MODERATOR','CONTENT_ADMIN'). Trang quản trị "
                 "/admin/safety-rules vốn chỉ dành cho SYSTEM_ADMIN nên backend rộng hơn frontend. "
                 "Ghi chú: enum Role không có giá trị ADMIN, nên 'ADMIN' trong annotation là giá trị chết.",
        "fix": "Siết về @PreAuthorize(\"hasRole('SYSTEM_ADMIN')\") cấp class. Chủ dự án xác nhận "
               "đây là phạm vi đúng.",
    },
    {
        "id": "DEF-RBAC-002",
        "match": ("AiModerationAdminControllerSecurityTest", "listPolicies_asNonSystemAdmin_returns403"),
        "severity": "Cao — RBAC",
        "component": "05_Development/CareBridgeAPI/.../aimoderation/controller/AiModerationAdminController.java",
        "symptom": "MODERATOR, CONTENT_ADMIN, EXPERT, MOTHER đọc được toàn bộ danh mục AI policy "
                   "của quản trị (kỳ vọng 403, thực tế 200).",
        "cause": "Cùng commit ff31c960 nới quyền cấp class cho controller quản trị AI moderation.",
        "fix": "Siết class về hasRole('SYSTEM_ADMIN'); riêng GET /status giữ "
               "hasAnyRole('SYSTEM_ADMIN','MODERATOR') vì trang PendingContentQueuePage của "
               "kiểm duyệt viên cần đọc trạng thái AI. CONTENT_ADMIN không nằm trong ngoại lệ này.",
    },
    {
        "id": "DEF-DATA-003",
        "match": ("growth_measurement_form_test.dart", "edit preserves a null note"),
        "severity": "Trung bình — sai dữ liệu",
        "component": "05_Development/CareBridgeMobileApp/lib/features/healthRecords/screens/"
                     "growth_measurement_form_screen.dart",
        "symptom": "Sửa một bản ghi tăng trưởng có ghi chú đang là null, người dùng không đụng vào "
                   "ô ghi chú, payload PATCH vẫn gửi note: \"\" và ghi đè null thành chuỗi rỗng.",
        "cause": "_noteController khởi tạo bằng existing?.note ?? '' nên khi đọc lại luôn ra chuỗi "
                 "rỗng, trong khi điều kiện so sánh lại đối chiếu với null thô → luôn khác nhau.",
        "fix": "Chuẩn hoá cả hai vế trước khi so sánh (existing?.note?.trim() ?? ''). Bổ sung test "
               "cho chiều ngược lại: xoá ghi chú đang có vẫn phải gửi note: \"\".",
    },
]

# Root-cause notes for the defects that are intentionally left red. The raw failure text alone does
# not say why the test must stay red, and that is the one thing a reviewer needs.
KNOWN_DEFECT_NOTES = {
    "allEndpoints_asModerator_shouldReturn403":
        "Commit ff31c960 nới @PreAuthorize cấp class của RedFlagRuleController thành "
        "hasAnyRole('SYSTEM_ADMIN','ADMIN','MODERATOR','CONTENT_ADMIN'), nên MODERATOR tạo/sửa/xoá "
        "được red-flag rule. /admin/safety-rules là trang của SYSTEM_ADMIN → nghi ngờ nới quyền "
        "ngoài ý muốn. GIỮ ĐỎ cho tới khi chủ dự án xác nhận phạm vi quyền.",
    "listPolicies_asModerator_returns403":
        "Cùng commit ff31c960 nới quyền cấp class cho AiModerationAdminController; MODERATOR đọc "
        "được toàn bộ AI policy của admin. GIỮ ĐỎ cho tới khi xác nhận đây là thay đổi cố ý.",
    "edit preserves a null note when it is unchanged":
        "GrowthMeasurementFormScreen khởi tạo _noteController = existing?.note ?? '' nên khi ghi "
        "chú gốc là null và người dùng không sửa, '' != null → payload PATCH vẫn gửi note: \"\". "
        "Lỗi thật (ghi đè null thành chuỗi rỗng), mức độ thấp; sửa 1 dòng ở tầng ứng dụng.",
}


# ---------------------------------------------------------------- group
def build_groups():
    rows = [r for r in json.load(open(f"{HERE}/all-tests.json")) if r["unit"]]
    pats = [(n, m, re.compile(p)) for n, m, p in USE_CASES]
    buckets = OrderedDict((n, {"module": m, "rows": []}) for n, m, _ in USE_CASES)
    for r in rows:
        for n, m, p in pats:
            if p.search(r["container"]):
                buckets[n]["rows"].append(r)
                break
        else:
            raise SystemExit(f"unmapped test container: {r['container']}")

    # One use case == one sheet. Splitting a use case across "(1)/(2)" tabs only made the tab bar
    # look like it held duplicates, so wide sheets are preferred over numbered fragments.
    sheets = []
    for name, data in buckets.items():
        if name in EXCLUDED_USE_CASES:
            continue
        items = sorted(data["rows"], key=lambda r: (r["layer"], r["container"], r["name"]))
        if not items:
            continue
        sheets.append({"use_case": name, "label": name, "module": data["module"], "rows": items})
    return sheets


def safe_title(label, used):
    t = re.sub(r"[\[\]:*?/\\]", "-", label)[:31].strip()
    base, n = t, 2
    while t.lower() in used:
        suffix = f"~{n}"
        t = base[:31 - len(suffix)] + suffix
        n += 1
    used.add(t.lower())
    return t


# ---------------------------------------------------------------- writers
def write_guideline(ws):
    ws.column_dimensions["A"].width = 130
    lines = [
        ("Hướng dẫn đọc tài liệu Unit Test — CareBridge (SEP490_G79)", True),
        ("", False),
        ("1. Tổng quan", True),
        (" - Cover: thông tin chung của tài liệu và lịch sử thay đổi.", False),
        (" - MethodList: danh sách use case đã được unit test, kèm module, sheet tương ứng và file test nguồn.", False),
        (" - Statistics: tổng hợp Passed / Failed / Untested và tỉ lệ N (Normal) / A (Abnormal) / B (Boundary).", False),
        (" - Defects: lỗi do unit test phát hiện — phần A là lỗi đã sửa trong đợt này, phần B là lỗi còn mở.", False),
        (" - Mỗi sheet còn lại là một use case: ma trận điều kiện — kỳ vọng — kết quả cho từng UTCID.", False),
        ("", False),
        ("2. Cách đọc một sheet use case", True),
        (" - Mỗi cột UTCIDxx là một test case (một test method của JUnit / flutter_test / Vitest).", False),
        (" - Khối Condition: 'Test target' = class/file test; 'Input condition' = tiền đề và dữ liệu đầu vào.", False),
        (" - Khối Confirm: 'Expected result' = kết quả kỳ vọng; 'Actual' = kết quả thực tế khi chạy.", False),
        (" - Khối Result: Type (N/A/B), Passed/Failed, ngày chạy, và Defect ID nếu test case đó từng", False),
        ("   phát hiện lỗi (kể cả lỗi đã sửa xong — xem sheet Defects).", False),
        (" - Dấu 'O' đánh dấu điều kiện/kỳ vọng đó thuộc về UTCID ở cột tương ứng.", False),
        ("", False),
        ("3. Mức chi tiết của các sheet tự sinh — đọc trước khi đối chiếu với báo cáo mẫu", True),
        (" - 78 sheet use case được SINH TỰ ĐỘNG từ kết quả chạy thật (surefire XML, flutter --machine,", False),
        ("   vitest --reporter=json). Các nguồn này chỉ cung cấp TÊN test case và trạng thái P/F/U;", False),
        ("   chúng không chứa giá trị tham số đầu vào hay giá trị trả về.", False),
        (" - Vì vậy 'Input condition' và 'Expected result' ở các sheet đó là Ý ĐỊNH của test case", False),
        ("   (diễn giải từ tên test), không phải giá trị dữ liệu cụ thể. Mỗi UTCID là một test case", False),
        ("   độc lập nên ma trận 'O' là một-đối-một, không phải tổ hợp giá trị.", False),
        (" - Muốn có giá trị dữ liệu thật như hai file Report5.1_..._reference*.xlsx thì phải dựng", False),
        ("   sheet thủ công từ mã nguồn test. Cách làm: 06_Testing/UnitTesting/guide.md, mục 7.", False),
        ("", False),
        ("4. Phân loại test case", True),
        (" - N (Normal): luồng hợp lệ, dữ liệu nằm trong miền giá trị bình thường.", False),
        (" - B (Boundary): giá trị biên (độ dài tối đa/tối thiểu, giới hạn phân trang, ngưỡng cấu hình...).", False),
        (" - A (Abnormal): dữ liệu không hợp lệ, thiếu quyền, không tìm thấy, xung đột, ném exception.", False),
        ("", False),
        ("5. Phạm vi", True),
        (" - Chỉ tính unit test chạy hoàn toàn trong bộ nhớ: service/policy/mapper/validator (Mockito),", False),
        ("   controller slice (@WebMvcTest), widget test Flutter, component/service test Vitest.", False),
        (" - Không tính các class *IntegrationTest / *PostgresTest / *EmbeddedPostgresTest / *SmokeTest:", False),
        ("   đây là integration test cần Docker hoặc PostgreSQL nhúng, thuộc báo cáo Integration/E2E.", False),
        (" - Không tính các nhóm hạ tầng xuyên suốt (nạp .env, exception handler, migration contract,", False),
        ("   job dọn dữ liệu, audit trail, adapter Firebase, search): chúng không tương ứng use case nào.", False),
        (" - Mỗi use case đúng một sheet — không tách thành nhiều tab (1)/(2).", False),
        (" - Bản đồ chi tiết use case ↔ file test: 06_Testing/UnitTesting/linkFileTest.md", False),
        ("", False),
        ("6. Cách chạy lại", True),
        (" - Backend : cd 05_Development/CareBridgeAPI      && ./mvnw test", False),
        (" - Mobile  : cd 05_Development/CareBridgeMobileApp && flutter test", False),
        (" - Web     : cd 05_Development/CareBridgeWebApp    && npm test", False),
    ]
    for i, (text, bold) in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=text)
        if bold:
            c.font = B_BOLD


def write_cover(ws):
    for col, w in zip("ABCDEF", (26, 26, 22, 14, 34, 30)):
        ws.column_dimensions[col].width = w
    ws.merge_cells("B2:E2")
    t = ws["B2"]
    t.value = "UNIT TEST DOCUMENT"
    t.font = Font(bold=True, size=18, color="FFFFFF")
    t.fill = TITLE_FILL
    t.alignment = CENTER
    ws.row_dimensions[2].height = 30

    pairs = [
        (4, "Project Name", PROJECT_NAME, "Creator", CREATOR),
        (5, "Project Code", PROJECT_CODE, "Issue Date", ISSUE_DATE),
        (6, "Document Code", DOC_CODE, "Version", "1.1"),
        (7, "Repository", "CareBridge_SEP490_G79 (branch HuyND)", "Reviewer/Approver", ""),
    ]
    for r, k1, v1, k2, v2 in pairs:
        ws.cell(row=r, column=1, value=k1).font = B_BOLD
        ws.cell(row=r, column=2, value=v1)
        ws.cell(row=r, column=5, value=k2).font = B_BOLD
        ws.cell(row=r, column=6, value=v2)

    ws.cell(row=9, column=1, value="Record of change").font = B_BOLD
    head = ["Effective Date", "Version", "Change Item", "*A,D,M", "Change description", "Reference"]
    for i, h in enumerate(head, 1):
        c = ws.cell(row=10, column=i, value=h)
        c.font = B_BOLD
        c.fill = HDR_FILL
        c.border = BOX
    changes = [
        [ISSUE_DATE, "1.0", "All sheets", "A",
         "Khởi tạo báo cáo Unit Test từ kết quả chạy thật của 3 tầng Backend / Mobile / Web",
         "06_Testing/UnitTesting/linkFileTest.md"],
        [REVISION_DATE, "1.1", "Guideline, Defects, sheet use case", "M",
         "Bổ sung sheet Defects (3 lỗi do unit test phát hiện và đã sửa) và sheet mẫu "
         "'MAU data-driven (demo)'; ô 'Test requirement' đổi từ danh sách file test sang phát biểu "
         "yêu cầu; siết lại cách phân loại Boundary (loại các ranh giới kiến trúc bị bắt nhầm).",
         "Đối chiếu Report5.1_UnitTest_reference1/2.xlsx"],
    ]
    for r, row in enumerate(changes, 11):
        for i, v in enumerate(row, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = BOX
            c.alignment = WRAP


def write_method_list(ws, sheets, titles):
    for col, w in zip("ABCDE", (6, 22, 34, 34, 80)):
        ws.column_dimensions[col].width = w
    ws.cell(row=2, column=1, value="Method / Use Case List").font = Font(bold=True, size=14)
    ws.cell(row=4, column=1, value="Project Name").font = B_BOLD
    ws.cell(row=4, column=3, value=PROJECT_NAME)
    ws.cell(row=5, column=1, value="Project Code").font = B_BOLD
    ws.cell(row=5, column=3, value=PROJECT_CODE)
    ws.cell(row=6, column=1, value="Test Environment").font = B_BOLD
    ws.cell(row=6, column=3, value=(
        "1. Java 21 + Spring Boot (JUnit 5, Mockito, H2 MODE=PostgreSQL)\n"
        "2. Flutter SDK (flutter_test)\n"
        "3. Node.js + Vitest + React Testing Library (jsdom)"))
    ws.cell(row=6, column=3).alignment = WRAP
    ws.row_dimensions[6].height = 48

    head = ["No", "Module", "Use Case", "Sheet Name", "Source test files"]
    for i, h in enumerate(head, 1):
        c = ws.cell(row=8, column=i, value=h)
        c.font = B_BOLD
        c.fill = HDR_FILL
        c.border = BOX
        c.alignment = CENTER

    prev_module = None
    for n, (sh, title) in enumerate(zip(sheets, titles), 1):
        r = 8 + n
        files = sorted({x["container"] for x in sh["rows"]})
        ws.cell(row=r, column=1, value=n).border = BOX
        ws.cell(row=r, column=2, value="" if sh["module"] == prev_module else sh["module"]).border = BOX
        prev_module = sh["module"]
        ws.cell(row=r, column=3, value=sh["label"]).border = BOX
        link = ws.cell(row=r, column=4, value=title)
        link.hyperlink = f"#'{title}'!A1"
        link.style = "Hyperlink"
        link.border = BOX
        c = ws.cell(row=r, column=5, value="\n".join(files))
        c.alignment = WRAP
        c.border = BOX


def write_statistics(ws, sheets, titles, stats):
    for col, w in zip("ABCDEFGHIJ", (5, 34, 11, 10, 11, 8, 8, 8, 15, 10)):
        ws.column_dimensions[col].width = w
    ws.merge_cells("B2:H2")
    t = ws["B2"]
    t.value = "UNIT TEST REPORT"
    t.font = Font(bold=True, size=16, color="FFFFFF")
    t.fill = TITLE_FILL
    t.alignment = CENTER
    ws.row_dimensions[2].height = 26

    meta = [(4, "Project Name", PROJECT_NAME, "Creator", CREATOR),
            (5, "Project Code", PROJECT_CODE, "Reviewer/Approver", ""),
            (6, "Document Code", DOC_CODE, "Issue Date", ISSUE_DATE)]
    for r, k1, v1, k2, v2 in meta:
        ws.cell(row=r, column=2, value=k1).font = B_BOLD
        ws.cell(row=r, column=3, value=v1)
        ws.cell(row=r, column=6, value=k2).font = B_BOLD
        ws.cell(row=r, column=8, value=v2)
    ws.cell(row=7, column=2, value="Notes").font = B_BOLD
    ws.cell(row=7, column=3, value=(
        "Số liệu lấy trực tiếp từ lần chạy thật: ./mvnw test (backend), flutter test (mobile), "
        "npm test (web). Không tính *IntegrationTest / *PostgresTest / *EmbeddedPostgresTest — "
        "các class này cần Docker/PostgreSQL nhúng và thuộc báo cáo Integration."))
    ws.cell(row=7, column=3).alignment = WRAP
    ws.row_dimensions[7].height = 34

    head = ["No", "Function code (Use case)", "Passed", "Failed", "Untested", "N", "A", "B", "Total Test Cases"]
    for i, h in enumerate(head, 2):
        c = ws.cell(row=10, column=i, value=h)
        c.font = B_BOLD
        c.fill = HDR_FILL
        c.border = BOX
        c.alignment = CENTER

    r = 11
    for n, (sh, title, st) in enumerate(zip(sheets, titles, stats), 1):
        ws.cell(row=r, column=2, value=n).border = BOX
        link = ws.cell(row=r, column=3, value=sh["label"])
        link.hyperlink = f"#'{title}'!A1"
        link.style = "Hyperlink"
        link.border = BOX
        for i, key in enumerate(["passed", "failed", "untested", "N", "A", "B", "total"]):
            c = ws.cell(row=r, column=4 + i, value=st[key])
            c.border = BOX
            c.alignment = CENTER
        if st["failed"]:
            ws.cell(row=r, column=5).fill = FAIL_FILL
        r += 1

    sub = r + 1
    ws.cell(row=sub, column=3, value="Sub total").font = B_BOLD
    for i, key in enumerate(["passed", "failed", "untested", "N", "A", "B", "total"]):
        c = ws.cell(row=sub, column=4 + i, value=sum(s[key] for s in stats))
        c.font = B_BOLD
        c.fill = HDR_FILL
        c.border = BOX
        c.alignment = CENTER

    total = sum(s["total"] for s in stats) or 1
    executed = sum(s["passed"] + s["failed"] for s in stats)
    metrics = [
        ("Test coverage (đã thực thi / tổng)", executed / total),
        ("Test successful coverage (Passed / tổng)", sum(s["passed"] for s in stats) / total),
        ("Normal case", sum(s["N"] for s in stats) / total),
        ("Abnormal case", sum(s["A"] for s in stats) / total),
        ("Boundary case", sum(s["B"] for s in stats) / total),
    ]
    rr = sub + 2
    for label, val in metrics:
        ws.cell(row=rr, column=3, value=label).font = B_BOLD
        c = ws.cell(row=rr, column=5, value=round(val * 100, 2))
        c.number_format = "0.00"
        ws.cell(row=rr, column=6, value="%")
        rr += 1
    ws.freeze_panes = "A11"


def write_use_case_sheet(ws, sheet, defect_prefix):
    items = sheet["rows"]
    n = len(items)
    first_col = 6  # column F

    parsed = []
    for row in items:
        target, cond, exp = split_case(row)
        parsed.append({"row": row, "target": target, "cond": cond, "exp": exp,
                       "type": case_type(row, cond, exp)})

    stats = {"passed": sum(1 for p in parsed if p["row"]["status"] == "passed"),
             "failed": sum(1 for p in parsed if p["row"]["status"] == "failed"),
             "untested": sum(1 for p in parsed if p["row"]["status"] == "untested"),
             "N": sum(1 for p in parsed if p["type"] == "N"),
             "A": sum(1 for p in parsed if p["type"] == "A"),
             "B": sum(1 for p in parsed if p["type"] == "B"),
             "total": n}
    assert stats["N"] + stats["A"] + stats["B"] == n
    assert stats["passed"] + stats["failed"] + stats["untested"] == n

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 62
    ws.column_dimensions["E"].width = 3
    for i in range(n):
        ws.column_dimensions[get_column_letter(first_col + i)].width = 9

    def head(r, k1, v1, k2=None, v2=None):
        ws.cell(row=r, column=1, value=k1).font = B_BOLD
        ws.cell(row=r, column=3, value=v1)
        if k2:
            ws.cell(row=r, column=6, value=k2).font = B_BOLD
            ws.cell(row=r, column=12, value=v2)

    head(1, "Code Module", sheet["module"], "Use Case", sheet["label"])
    head(2, "Created By", CREATOR, "Executed By", CREATOR)
    # The requirement statement, not the source-file list: the files are already in MethodList
    # column E and in 06_Testing/UnitTesting/linkFileTest.md.
    ws.cell(row=3, column=1, value="Test requirement").font = B_BOLD
    c = ws.cell(row=3, column=3, value=REQUIREMENTS[sheet["use_case"]])
    c.alignment = WRAP
    ws.row_dimensions[3].height = 30

    for col, label in ((1, "Passed"), (3, "Failed"), (6, "Untested"), (12, "N/A/B"), (15, "Total Test Cases")):
        ws.cell(row=4, column=col, value=label).font = B_BOLD
    ws.cell(row=5, column=1, value=stats["passed"]).fill = PASS_FILL
    failed_cell = ws.cell(row=5, column=3, value=stats["failed"])
    if stats["failed"]:
        failed_cell.fill = FAIL_FILL
    ws.cell(row=5, column=6, value=stats["untested"])
    ws.cell(row=5, column=12, value=stats["N"])
    ws.cell(row=5, column=13, value=stats["A"])
    ws.cell(row=5, column=14, value=stats["B"])
    ws.cell(row=5, column=15, value=stats["total"]).font = B_BOLD

    ids = [f"UTCID{i:02d}" for i in range(1, n + 1)]
    for i, uid in enumerate(ids):
        c = ws.cell(row=7, column=first_col + i, value=uid)
        c.font = B_BOLD
        c.fill = HDR_FILL
        c.border = BOX
        c.alignment = Alignment(horizontal="center", vertical="bottom", text_rotation=90)
    ws.row_dimensions[7].height = 62

    r = 8

    def section(label):
        nonlocal r
        c = ws.cell(row=r, column=1, value=label)
        c.font = B_BOLD
        c.fill = SEC_FILL
        return c

    def group(label):
        nonlocal r
        ws.cell(row=r, column=2, value=label).font = B_BOLD
        ws.cell(row=r, column=2).fill = SEC_FILL
        r += 1

    def marked_row(text, cols):
        nonlocal r
        c = ws.cell(row=r, column=4, value=text)
        c.alignment = WRAP
        c.border = BOX
        for i in range(n):
            cc = ws.cell(row=r, column=first_col + i)
            cc.border = BOX
            cc.alignment = CENTER
            if i in cols:
                cc.value = "O"
        r += 1

    section("Condition")
    group("Precondition")
    marked_row("Các dependency được mock/stub; test chạy trong bộ nhớ, không phụ thuộc DB thật hay mạng ngoài",
               set(range(n)))
    group("Test target")
    by_target = OrderedDict()
    for i, p in enumerate(parsed):
        by_target.setdefault(p["target"], []).append(i)
    for target, cols in by_target.items():
        marked_row(target, set(cols))
    group("Input condition")
    by_cond = OrderedDict()
    for i, p in enumerate(parsed):
        by_cond.setdefault(p["cond"], []).append(i)
    for cond, cols in by_cond.items():
        marked_row(cond, set(cols))

    section("Confirm")
    group("Expected result")
    by_exp = OrderedDict()
    for i, p in enumerate(parsed):
        by_exp.setdefault(p["exp"], []).append(i)
    for exp, cols in by_exp.items():
        marked_row(exp, set(cols))
    group("Actual")
    marked_row("Đúng kỳ vọng", {i for i, p in enumerate(parsed) if p["row"]["status"] == "passed"})
    marked_row("Sai kỳ vọng — xem Defect ID", {i for i, p in enumerate(parsed) if p["row"]["status"] == "failed"})
    marked_row("Không thực thi (skipped)", {i for i, p in enumerate(parsed) if p["row"]["status"] == "untested"})

    section("Result")
    ws.cell(row=r, column=2, value="Type (N: Normal, A: Abnormal, B: Boundary)").font = B_BOLD
    for i, p in enumerate(parsed):
        c = ws.cell(row=r, column=first_col + i, value=p["type"])
        c.alignment = CENTER
        c.border = BOX
    r += 1
    ws.cell(row=r, column=2, value="Passed/Failed").font = B_BOLD
    # A test that caught a defect keeps the defect number even though it is green again — that link
    # is the evidence the case did its job, and it is what the Defects sheet cross-references.
    defects = {}
    fixed_here = []
    for i, p in enumerate(parsed):
        for fd in FIXED_DEFECTS:
            cls, prefix = fd["match"]
            if p["row"]["short"] == cls and p["row"]["name"].startswith(prefix):
                defects[i] = fd["id"]
                if fd not in fixed_here:
                    fixed_here.append(fd)
    dn = 0
    open_defects = []
    for i, p in enumerate(parsed):
        st = p["row"]["status"]
        v = {"passed": "P", "failed": "F", "untested": "U"}[st]
        c = ws.cell(row=r, column=first_col + i, value=v)
        c.alignment = CENTER
        c.border = BOX
        c.fill = {"P": PASS_FILL, "F": FAIL_FILL, "U": UNT_FILL}[v]
        if st == "failed":
            dn += 1
            defects[i] = f"{defect_prefix}-{dn:02d}"
            open_defects.append((defects[i], p))
    r += 1
    ws.cell(row=r, column=2, value="Executed Date").font = B_BOLD
    for i in range(n):
        c = ws.cell(row=r, column=first_col + i, value=ISSUE_DATE)
        c.alignment = CENTER
        c.border = BOX
    r += 1
    ws.cell(row=r, column=2, value="Defect ID").font = B_BOLD
    for i in range(n):
        c = ws.cell(row=r, column=first_col + i, value=defects.get(i, ""))
        c.alignment = CENTER
        c.border = BOX

    ws.freeze_panes = ws.cell(row=8, column=first_col)
    return stats, open_defects, fixed_here


# ---------------------------------------------------------------- data-driven sample
# Off by default: the workbook ships only the auto-generated sheets. Flip to True to emit the
# hand-built reference-style sheet — it is kept as the template to copy when a use case is
# converted to real input/output values. See 06_Testing/UnitTesting/guide.md, mục 7.
EMIT_SAMPLE_SHEET = False

# The auto-generated sheets can only print what surefire/flutter/vitest emit: a test name and a
# status. The reference reports print the *values* — `username_test`, `null`, `NotFoundException`,
# the literal log message — because they were written by hand from the test source.
# This sheet is one use case rebuilt that way, from RedFlagRuleControllerSecurityTest, so the two
# styles can be compared side by side before deciding whether to convert the rest.
# Column B holds the parameter name (as in the references), column D the concrete value.
SAMPLE = {
    "title": "MAU data-driven (demo)",
    "module": "Triage / Safety Rules",
    "use_case": "Triage Red Flag Rules — quản trị quy tắc dấu hiệu nguy hiểm",
    "requirement": REQUIREMENTS["Triage Red Flag Rules"],
    "source": "CareBridgeAPI/src/test/java/com/carebridge/backend/security/"
              "RedFlagRuleControllerSecurityTest.java",
    "n": 7,
    "precondition": [
        ("RedFlagRuleService được mock; SecurityConfig thật được nạp qua @Import", "1234567"),
        ("Bảng red_flag_rules có sẵn rule id 00000000-0000-0000-0000-000000000002", "1234567"),
    ],
    "sections": [
        ("Condition", [
            ("Vai trò trong JWT", [
                ("MODERATOR", "1"),
                ("CONTENT_ADMIN", "2"),
                ("EXPERT", "3"),
                ("MOTHER", "4"),
                ("FAMILY", "5"),
                ("SYSTEM_ADMIN", "6"),
                ("(không gửi Authorization header)", "7"),
            ]),
            ("userId trong token", [
                ("00000000-0000-0000-0000-0000000000bb", "12345"),
                ("00000000-0000-0000-0000-0000000000aa", "6"),
                ("(không có)", "7"),
            ]),
            ("Endpoint được gọi", [
                ("POST   /api/v1/admin/red-flag-rules", "12345"),
                ("GET    /api/v1/admin/red-flag-rules", "1234567"),
                ("PATCH  /api/v1/admin/red-flag-rules/{ruleId}", "12345"),
                ("DELETE /api/v1/admin/red-flag-rules/{ruleId}", "12345"),
            ]),
            ("Request body", [
                ('{"keyword":"tu khoa moi","severity":"RED","action":"ESCALATE"}  (POST)', "12345"),
                ('{"isActive":false}  (PATCH)', "12345"),
                ("(không có body)", "1234567"),
            ]),
            ("CSRF token", [
                ("có (.with(csrf()) cho POST/PATCH/DELETE)", "12345"),
                ("không áp dụng (chỉ gọi GET)", "67"),
            ]),
        ]),
        ("Confirm", [
            ("HTTP status", [
                ("403 Forbidden", "12345"),
                ("200 OK", "6"),
                ("401 Unauthorized", "7"),
            ]),
            ("Response body", [
                ("(rỗng)", "1234"),
                ('{"items":[],"totalElements":0,"page":0,"size":20}', "6"),
                ("(rỗng — HttpStatusEntryPoint không ghi body)", "7"),
            ]),
            ("Tương tác với RedFlagRuleService", [
                ("never(): createRule / listRules / updateRule / deleteRule", "12345"),
                ("listRules(any()) được gọi đúng 1 lần", "6"),
                ("không gọi — bị chặn ở filter chain trước khi vào controller", "7"),
            ]),
            ("Exception", [
                ("AccessDeniedException (Spring Security xử lý → 403)", "12345"),
                ("(không có)", "6"),
                ("AuthenticationException (→ 401)", "7"),
            ]),
            ("Actual", [
                ("Đúng kỳ vọng", "1234567"),
            ]),
        ]),
    ],
    "types": "AAAAANA",
    "verdicts": "PPPPPPP",
    "defect_ids": ["DEF-RBAC-001"] * 5 + ["", ""],
}


def write_sample_sheet(ws):
    s = SAMPLE
    n = s["n"]
    first_col = 6
    for col, w in zip("ABCDE", (12, 34, 3, 66, 3)):
        ws.column_dimensions[col].width = w
    for i in range(n):
        ws.column_dimensions[get_column_letter(first_col + i)].width = 9

    ws.cell(row=1, column=1, value="Code Module").font = B_BOLD
    ws.cell(row=1, column=3, value=s["module"])
    ws.cell(row=1, column=6, value="Use Case").font = B_BOLD
    ws.cell(row=1, column=12, value=s["use_case"])
    ws.cell(row=2, column=1, value="Created By").font = B_BOLD
    ws.cell(row=2, column=3, value=CREATOR)
    ws.cell(row=2, column=6, value="Executed By").font = B_BOLD
    ws.cell(row=2, column=12, value=CREATOR)
    ws.cell(row=3, column=1, value="Test requirement").font = B_BOLD
    ws.cell(row=3, column=3, value=s["requirement"]).alignment = WRAP
    ws.row_dimensions[3].height = 30

    note = ws.cell(row=5, column=1, value=(
        "SHEET MẪU — dựng thủ công từ mã nguồn test (" + s["source"] + "). "
        "Khác các sheet tự sinh ở chỗ: cột D ghi GIÁ TRỊ đầu vào/đầu ra thật, cột B ghi TÊN THAM SỐ, "
        "và một UTCID là một tổ hợp giá trị — đúng cách hai file reference đang làm."))
    note.font = Font(bold=True, color="9C4500")
    note.alignment = WRAP
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=first_col + n - 1)
    ws.row_dimensions[5].height = 34

    ids = [f"UTCID{i:02d}" for i in range(1, n + 1)]
    for i, uid in enumerate(ids):
        c = ws.cell(row=7, column=first_col + i, value=uid)
        c.font = B_BOLD
        c.fill = HDR_FILL
        c.border = BOX
        c.alignment = Alignment(horizontal="center", vertical="bottom", text_rotation=90)
    ws.row_dimensions[7].height = 62

    r = 8

    def value_row(text, cols):
        nonlocal r
        c = ws.cell(row=r, column=4, value=text)
        c.alignment = WRAP
        c.border = BOX
        for i in range(n):
            cc = ws.cell(row=r, column=first_col + i)
            cc.border = BOX
            cc.alignment = CENTER
            if str(i + 1) in cols:
                cc.value = "O"
        r += 1

    ws.cell(row=r, column=1, value="Condition").font = B_BOLD
    ws.cell(row=r, column=1).fill = SEC_FILL
    ws.cell(row=r, column=2, value="Precondition").font = B_BOLD
    ws.cell(row=r, column=2).fill = SEC_FILL
    r += 1
    for text, cols in s["precondition"]:
        value_row(text, cols)

    for section, groups in s["sections"]:
        if section != "Condition":
            c = ws.cell(row=r, column=1, value=section)
            c.font = B_BOLD
            c.fill = SEC_FILL
        for param, values in groups:
            ws.cell(row=r, column=2, value=param).font = B_BOLD
            ws.cell(row=r, column=2).fill = SEC_FILL
            r += 1
            for text, cols in values:
                value_row(text, cols)

    c = ws.cell(row=r, column=1, value="Result")
    c.font = B_BOLD
    c.fill = SEC_FILL
    ws.cell(row=r, column=2, value="Type (N: Normal, A: Abnormal, B: Boundary)").font = B_BOLD
    for i in range(n):
        cc = ws.cell(row=r, column=first_col + i, value=s["types"][i])
        cc.alignment = CENTER
        cc.border = BOX
    r += 1
    ws.cell(row=r, column=2, value="Passed/Failed").font = B_BOLD
    for i in range(n):
        cc = ws.cell(row=r, column=first_col + i, value=s["verdicts"][i])
        cc.alignment = CENTER
        cc.border = BOX
        cc.fill = PASS_FILL if s["verdicts"][i] == "P" else FAIL_FILL
    r += 1
    ws.cell(row=r, column=2, value="Executed Date").font = B_BOLD
    for i in range(n):
        cc = ws.cell(row=r, column=first_col + i, value=ISSUE_DATE)
        cc.alignment = CENTER
        cc.border = BOX
    r += 1
    ws.cell(row=r, column=2, value="Defect ID").font = B_BOLD
    for i in range(n):
        cc = ws.cell(row=r, column=first_col + i, value=s["defect_ids"][i])
        cc.alignment = CENTER
        cc.border = BOX
    ws.freeze_panes = ws.cell(row=8, column=first_col)


def write_defects(ws, open_defects, fixed_defects):
    for col, w in zip("ABCDEFGH", (15, 26, 17, 46, 58, 58, 58, 20)):
        ws.column_dimensions[col].width = w
    ws.cell(row=1, column=1, value="DEFECT LIST — lỗi do unit test phát hiện").font = Font(bold=True, size=14)

    def header(r, title, cols):
        c = ws.cell(row=r, column=1, value=title)
        c.font = Font(bold=True, size=12)
        for i, h in enumerate(cols, 1):
            hc = ws.cell(row=r + 1, column=i, value=h)
            hc.font = B_BOLD
            hc.fill = HDR_FILL
            hc.border = BOX
            hc.alignment = CENTER
        return r + 2

    r = header(2, "A. Đã sửa trong đợt test này (test hiện xanh; revert bản sửa thì case tương ứng đỏ lại)",
               ["Defect ID", "Use case (sheet)", "Mức độ", "Thành phần lỗi",
                "Hiện tượng", "Nguyên nhân gốc", "Cách sửa", "Trạng thái"])
    for fd in fixed_defects:
        vals = [fd["id"], fd["sheet"], fd["severity"], fd["component"],
                fd["symptom"], fd["cause"], fd["fix"], "Đã sửa — đã verify lại"]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = BOX
            c.alignment = WRAP
        ws.cell(row=r, column=8).fill = PASS_FILL
        ws.row_dimensions[r].height = 78
        r += 1
    if not fixed_defects:
        ws.cell(row=r, column=1, value="(không có)")
        r += 1

    r += 2
    r = header(r, "B. Còn mở — unit test đang FAIL tại thời điểm phát hành báo cáo",
               ["Defect ID", "Use case (sheet)", "Test class / file", "Test case",
                "Nhận định / nguyên nhân", "Thông điệp lỗi thô", "Type"])
    for did, label, p in open_defects:
        vals = [did, label, p["row"]["container"], p["row"]["name"],
                KNOWN_DEFECT_NOTES.get(p["row"]["name"], ""),
                re.sub(r"\s+", " ", p["row"].get("detail", ""))[:300], p["type"]]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = BOX
            c.alignment = WRAP
        r += 1
    if not open_defects:
        c = ws.cell(row=r, column=1, value="Không còn lỗi mở: 0 test FAIL trên cả ba tầng "
                                           "(backend / mobile / web) ở lần chạy cuối.")
        c.fill = PASS_FILL


def main():
    sheets = build_groups()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    write_guideline(wb.create_sheet("Guideline"))
    write_cover(wb.create_sheet("Cover"))
    ws_methods = wb.create_sheet("MethodList")
    ws_stats = wb.create_sheet("Statistics")
    ws_def = wb.create_sheet("Defects")
    if EMIT_SAMPLE_SHEET:
        write_sample_sheet(wb.create_sheet(SAMPLE["title"]))

    used, titles, stats, open_defects, fixed_defects = set(), [], [], [], []
    for idx, sh in enumerate(sheets, 1):
        title = safe_title(sh["label"], used)
        titles.append(title)
        ws = wb.create_sheet(title)
        st, defs, fixed = write_use_case_sheet(ws, sh, f"DEF-{idx:03d}")
        stats.append(st)
        for did, p in defs:
            open_defects.append((did, sh["label"], p))
        for fd in fixed:
            fixed_defects.append({**fd, "sheet": sh["label"]})

    # Every declared defect must have landed on a sheet; a typo in `match` would otherwise drop it
    # silently and put the report right back to claiming zero defects.
    missing = {fd["id"] for fd in FIXED_DEFECTS} - {fd["id"] for fd in fixed_defects}
    assert not missing, f"FIXED_DEFECTS not matched to any test case: {missing}"

    write_method_list(ws_methods, sheets, titles)
    write_statistics(ws_stats, sheets, titles, stats)
    write_defects(ws_def, open_defects, fixed_defects)

    grand = {k: sum(s[k] for s in stats) for k in ("passed", "failed", "untested", "N", "A", "B", "total")}
    assert grand["N"] + grand["A"] + grand["B"] == grand["total"]
    assert grand["passed"] + grand["failed"] + grand["untested"] == grand["total"]

    wb.save(OUT)
    print(f"sheets={len(sheets)} open_defects={len(open_defects)} fixed_defects={len(fixed_defects)}")
    print("grand:", grand)
    print("saved:", OUT)


if __name__ == "__main__":
    main()
