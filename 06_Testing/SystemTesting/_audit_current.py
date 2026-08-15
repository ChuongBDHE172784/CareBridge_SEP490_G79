import openpyxl,json,sys,os,re
p=sys.argv[1]
w=openpyxl.load_workbook(p,data_only=False)
rows=[]; fonts={}; ev=0
for s in w.worksheets:
 if s.title in ('Cover','Test Cases','Test Statistics'): continue
 for r in range(1,s.max_row+1):
  i=s.cell(r,1).value
  if isinstance(i,str) and i.startswith('ST-'):
   rr={'sheet':s.title,'row':r,'id':i,'obj':s.cell(r,2).value,'ev':s.cell(r,6).value,'r1':s.cell(r,7).value,'r2':s.cell(r,10).value,'r3':s.cell(r,13).value,'note':s.cell(r,16).value}
   rows.append(rr)
   ev+=1 if rr['ev'] else 0
 for row in s.iter_rows():
  for c in row:
   if c.value is not None:
    fonts[c.font.name or 'None']=fonts.get(c.font.name or 'None',0)+1
counts={}
for k in ('r1','r2','r3'):
 counts[k]={z:sum(1 for x in rows if x[k]==z) for z in ('Passed','Failed','Pending')}
print(json.dumps({'rows':rows,'counts':counts,'evidence':ev,'fonts':fonts},ensure_ascii=False,default=str))