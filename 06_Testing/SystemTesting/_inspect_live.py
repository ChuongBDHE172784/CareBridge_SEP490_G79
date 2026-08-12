import openpyxl,json,sys
p=sys.argv[1]
w=openpyxl.load_workbook(p,data_only=False)
out={}
for s in w.worksheets:
 if s.title=='Test Statistics': continue
 rows=[]
 for r in range(1,s.max_row+1):
  v=s.cell(r,1).value
  if isinstance(v,str) and v.startswith('ST-'):
   rows.append({'row':r,'id':v,'obj':s.cell(r,2).value,'r1':s.cell(r,7).value,'r2':s.cell(r,10).value,'r3':s.cell(r,13).value,'ev':bool(s.cell(r,6).value)})
 if rows: out[s.title]=rows
print(json.dumps(out,ensure_ascii=False,default=str))