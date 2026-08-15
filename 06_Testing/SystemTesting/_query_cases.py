import openpyxl,json,sys
w=openpyxl.load_workbook(sys.argv[1],data_only=False)
ids=set(sys.argv[2].split(','))
out=[]
for s in w.worksheets:
 for r in range(1,s.max_row+1):
  if str(s.cell(r,1).value) in ids:
   out.append({'sheet':s.title,'row':r,'id':s.cell(r,1).value,'obj':s.cell(r,2).value,'ev':s.cell(r,6).value,'r1':s.cell(r,7).value,'r2':s.cell(r,10).value,'r3':s.cell(r,13).value})
print(json.dumps(out,ensure_ascii=False,default=str))